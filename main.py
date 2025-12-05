import streamlit as st
from supabase import create_client, Client
import datetime
import json
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

# --- 1. SETUP SUPABASE ---
try:
    url = st.secrets["supabase"]["url"]
    key = st.secrets["supabase"]["key"]
    supabase: Client = create_client(url, key)
except Exception as e:
    st.error("Setup Supabase gagal. Pastikan file .streamlit/secrets.toml sudah benar.")
    st.stop()

# --- 2. DATA REFERENSI ---
LIST_UNIT_KERJA = [
    "DDK", "DGT", "DIK", "DIT", "DJK", "DMM", "DMP", "DMU", "DOC", "DPI", 
    "DST", "DTC", "IPR", "KJF", "KPS", "KRH", "KRR", "KRS", "KRU", "KSI", "KST",
    "Balai Besar Meteorologi Klimatologi dan Geofisika Wil. I – Medan",
    "Balai Besar Meteorologi Klimatologi dan Geofisika Wil. II – Tangerang Selatan",
    "Balai Besar Meteorologi Klimatologi dan Geofisika Wil. III – Badung",
    "Balai Besar Meteorologi Klimatologi dan Geofisika Wil. IV – Makassar",
    "Balai Besar Meteorologi Klimatologi dan Geofisika Wil. IV – Jayapura",
    "Stasiun Geofisika Kelas I Padang Panjang",
    "Stasiun Meteorologi Kelas I Kualanamu - Deli Serdang",
    "Stasiun Klimatologi Kelas IV Riau",
    "Stasiun Meteorologi Kelas II Maritim Belawan – Medan",
    "Stasiun Klimatologi Kelas I Jawa Barat",
    "Stasiun Meteorologi Kelas II Maritim Tanjung Emas – Semarang",
    "Stasiun Geofisika Kelas I Sleman",
    "Stasiun Meteorologi Kelas I Soekarno Hatta – Tangerang",
    "Stasiun Geofisika Kelas II Pasuruan",
    "Stasiun Meteorologi Kelas I Juanda – Sidoarjo",
    "Stasiun Klimatologi Kelas I Kalimantan Selatan",
    "Stasiun Meteorologi Kelas II Maritim Tanjung Perak - Surabaya",
    "Stasiun Geofisika Kelas II Denpasar",
    "Stasiun Meteorologi Kelas II Maritim Bitung",
    "Stasiun Geofisika Kelas I Ambon",
    "Stasiun Klimatologi Kelas IV Gorontalo",
    "Stasiun Meteorologi Kelas I Sultan Hasanuddin – Makassar",
    "Stasiun Meteorologi Kelas III Mopah – Merauke",
    "Stasiun Meteorologi Kelas I Domine Eduard Osok – Sorong",
    "Stasiun Meteorologi Kelas I Sentani – Jayapura",
    "Stasiun Pemantau Atmosfer Global Puncak Vihara Klademak - Sorong"
]

PERTANYAAN_STORAGE = [
    "1. Jika Storage dikelola secara mandiri, jelaskan kapasitas keseluruhan total storage yang tersedia, dan yang tersisa saat ini (raw dan usable) ?",
    "2. Berapa tingkat utilisasi rata-rata dan utilisasi puncak data dalam 6–12 bulan terakhir?",
    "3. Berapa kapasitas cadangan yang tersedia untuk proyeksi pertumbuhan dari 1–3 tahun terakhir?",
    "4. Jenis dan arsitektur storage existing apa yang digunakan? (SAN/NAS/Object/Cloud)",
    "5. Berapa umur existing perangkat storage dan apakah masih berada dalam masa garansi/dukungan vendor?",
    "6. Data apa saja yang ingin disimpan (file/gambar/video)",
    "7. Apa saja tipe data yang diolah dan dihasilkan (Data Operational, Data Transactional, Data Analytic, Data Realtime atau Data Library/Archive)?",
    "8. Penyimpanan data apakah berdasarkan data kritikal seperti data satelit, radar, observasi, klimatologi dll",
    "9. Jika Database dikelola secara mandiri, berapa ukuran database existing yang ada di unit anda?",
    "10. Jika Database dikelola secara mandiri, adakah kendala terkait performance dari database?",
    "11. Perkiraan untuk masa retention backup existing berapa lama dan bagaimana justifikasinya?",
    "12. Apakah data sudah memiliki klasifikasi formal (Critical, Confidential, Internal, Public)",
    "13. Apakah sudah memiliki backup storage dan berapa besar kapasitas yang tersedia serta bagaimana utilisasinya?",
    "14. Jika belum memiliki backup storage, berapa kapasitas backup storage yang di inginkan?",
    "15. Jika sudah memiliki backup storage, bagaimana metode backup yang digunakan? (full backup, incremental backup, differential backup)",
    "16. Apakah dalam 5 tahun kedepan ada rencana penambahan kebutuhan storage?",
    "17. Catatan lain terkait storage yang ingin disampaikan"
]

PERTANYAAN_SERVER = [
    "1. Berapa jumlah server komputasi yang tersedia saat ini (fisik maupun VM)?",
    "2. Apa jenis platform yang digunakan? (Baremetal, Virtualisasi, HCI, Cloud, Hybrid)",
    "3. Berapa total kapasitas CPU (core), RAM dan Kapasitas Disk yang tersedia dan terpakai saat ini?",
    "4. Berapa rata-rata dan puncak utilisasi CPU, Memory dan Disk dalam 6–12 bulan terakhir?",
    "5. Aplikasi apa saja yang berjalan di server yang digunakan sesuai operasional di masing masing Unit Kerja? (Contoh, aplikasi Management Radar)",
    "6. Apakah saat ini ada kendala performa pada server (CPU bottleneck, memory leak, network throughput, disk I/O)?",
    "7. Apakah server memiliki Sistem Backup dan Redudansi?",
    "8. Apakah server saat ini berada dalam masa garansi maupun dukungan vendor?",
    "9. Berapa usia server yang saat ini digunakan?",
    "10. Apakah sudah ada standardisasi platform OS dan software environment?",
    "11. Apakah ditempat anda sudah ada alat monitoring server, jika iya bagaimana melakukannya ?",
    "12. Apakah ada rencana penambahan kapasitas server dalam 1–3 tahun ke depan?",
    "13. Apakah dalam 5 tahun kedepan ada rencana penambahan kebutuhan alat observasi?",
    "14. Catatan lain terkait server yang ingin disampaikan"
]

# --- 3. HELPER EXCEL STYLING ---
def create_styled_excel(data_storage, data_server, info):
    wb = Workbook()
    
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Light Blue
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=11)
    
    # --- SHEET 1: STORAGE ---
    ws1 = wb.active
    ws1.title = "Storage"
    setup_sheet(ws1, "Storage", data_storage, info, thin_border, header_fill, title_font, header_font, normal_font)

    # --- SHEET 2: SERVER KOMPUTASI ---
    ws2 = wb.create_sheet("Server Komputasi")
    setup_sheet(ws2, "Server Komputasi", data_server, info, thin_border, header_fill, title_font, header_font, normal_font)

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def setup_sheet(ws, type_name, data_list, info, border, fill, title_f, header_f, normal_f):
    # 1. Header Section (Logo & Title)
    # Merge cells for Title
    ws.merge_cells('A1:D5') 
    cell_title = ws['A1']
    cell_title.value = "Assessment Konsultansi Pembangunan Storage - BMKG"
    cell_title.font = title_f
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    cell_title.fill = fill
    
    # Insert Logo (BMKG)
    try:
        img = OpenpyxlImage('d:/Assesment_storage/logo-bmkg-png.png')
        img.height = 60
        img.width = 60
        # Adjust position slightly
        ws.add_image(img, 'A1')
    except Exception as e:
        # print(f"Logo not found: {e}")
        pass

    # 2. Metadata Section
    ws['A6'] = "Assessment Date :"
    ws['B6'] = info['date']
    ws['A7'] = "Unit Kerja (Eselon 2) :"
    ws['B7'] = info['unit_kerja']
    ws['A8'] = "PIC Unit Kerja (Eselon 2) :"
    ws['B8'] = info['pic_name']
    
    for row in range(6, 9):
        ws[f'A{row}'].font = header_f
        ws[f'B{row}'].font = normal_f
        # Merge B to D for cleaner look
        ws.merge_cells(f'B{row}:D{row}')

    # 3. Table Header
    headers = []
    if type_name == "Storage":
        headers = ["No", "Pertanyaan", "Keterangan", "Jawaban"]
        col_widths = [5, 60, 40, 40]
    else:
        headers = ["No", "Pertanyaan", "Skala Critical", "Jawaban"]
        col_widths = [5, 60, 20, 60]

    start_row = 10
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = header_f
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Grey for table header

    # 4. Table Data
    current_row = start_row + 1
    for item in data_list:
        # No
        c1 = ws.cell(row=current_row, column=1, value=item['No'])
        c1.border = border
        c1.alignment = Alignment(horizontal='center', vertical='top')
        
        # Pertanyaan
        c2 = ws.cell(row=current_row, column=2, value=item['Pertanyaan'])
        c2.border = border
        c2.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Col 3 (Keterangan / Skala)
        val_3 = item.get('Keterangan') if type_name == "Storage" else item.get('Skala Critical')
        c3 = ws.cell(row=current_row, column=3, value=val_3)
        c3.border = border
        c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Jawaban
        c4 = ws.cell(row=current_row, column=4, value=item['Jawaban'])
        c4.border = border
        c4.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        current_row += 1

    # Set Column Widths
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 5. Footer (Signatures)
    footer_start = current_row + 2
    
    # Headers
    ws.cell(row=footer_start, column=2, value="Diperiksa oleh").border = border
    ws.cell(row=footer_start, column=3, value="Disetujui oleh").border = border
    ws.cell(row=footer_start, column=4, value="Diketahui oleh").border = border
    
    # Center align headers
    for col in range(2, 5):
        cell = ws.cell(row=footer_start, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = header_f

    # Space for signature
    signature_height = 4
    for r in range(footer_start + 1, footer_start + 1 + signature_height):
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = border

    # Names
    name_row = footer_start + 1 + signature_height
    
    # Rachmat Kharisma
    c_rachmat = ws.cell(row=name_row, column=2, value="Rachmat Kharisma\n(PIC Assesment dan Pemetaan)")
    c_rachmat.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_rachmat.border = border
    c_rachmat.font = header_f
    
    # PIC Unit Kerja
    c_pic = ws.cell(row=name_row, column=3, value=f"{info['pic_name']}\n(PIC Unit Kerja)")
    c_pic.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_pic.border = border
    c_pic.font = header_f

    # Dr. Iqbal
    c_iqbal = ws.cell(row=name_row, column=4, value="Dr. Iqbal, S.Kom, M.T.I\nPIC Unit Data dan Komputasi")
    c_iqbal.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c_iqbal.border = border
    c_iqbal.font = header_f


# --- 4. UI LAYOUT ---
st.set_page_config(page_title="Assessment Storage BMKG", layout="wide")

# Header dengan Logo (Opsional jika ada URL logo)
st.title("Assessment Konsultansi Pembangunan Storage - BMKG")
st.markdown("---")

# Initialize session state for results
if 'assessment_results' not in st.session_state:
    st.session_state['assessment_results'] = None

# --- INFORMASI UMUM ---
st.subheader("I. Informasi Umum")
col1, col2, col3 = st.columns(3)
with col1:
    tgl_assessment = st.date_input("Assessment Date", datetime.date.today())
with col2:
    unit_kerja = st.selectbox("Unit Kerja (Eselon 2)", LIST_UNIT_KERJA)
with col3:
    pic_name = st.text_input("PIC Unit Kerja (Eselon 2)")

st.markdown("---")

# --- TAB NAVIGATION ---
tab1, tab2, tab3 = st.tabs(["💾 Storage Assessment", "🖥️ Server Komputasi Assessment", "📊 Hasil / Export"])

# Form Utama
with st.form("assessment_form"):
    
    with tab1:
        st.subheader("II. Storage Assessment")
        st.info("Silahkan isi Keterangan dan Jawaban untuk setiap pertanyaan.")
        
        storage_answers = {}
        
        h1, h2, h3 = st.columns([3, 2, 2])
        h1.markdown("**Pertanyaan**")
        h2.markdown("**Keterangan**")
        h3.markdown("**Jawaban**")
        
        for i, q in enumerate(PERTANYAAN_STORAGE):
            c1, c2, c3 = st.columns([3, 2, 2])
            with c1:
                st.write(q)
            with c2:
                ket = st.text_area(f"Ket_S_{i}", key=f"storage_ket_{i}", height=68, label_visibility="collapsed", placeholder="Keterangan...")
            with c3:
                ans = st.text_area(f"Jwb_S_{i}", key=f"storage_ans_{i}", height=68, label_visibility="collapsed", placeholder="Jawaban...")
                
            storage_answers[f"q_{i+1}"] = {
                "pertanyaan": q,
                "keterangan": ket,
                "jawaban": ans
            }
            st.divider()

    with tab2:
        st.subheader("III. Server Komputasi Assessment")
        st.info("Silahkan isi Skala Critical dan Jawaban untuk setiap pertanyaan.")

        server_answers = {}

        sh1, sh2, sh3 = st.columns([3, 1, 3])
        sh1.markdown("**Pertanyaan**")
        sh2.markdown("**Skala Critical**")
        sh3.markdown("**Jawaban**")

        skala_opsi = ["-", "Low", "Medium", "High", "Critical"]

        for i, q in enumerate(PERTANYAAN_SERVER):
            sc1, sc2, sc3 = st.columns([3, 1, 3])
            with sc1:
                st.write(q)
            with sc2:
                skala = st.selectbox(f"Skala_C_{i}", options=skala_opsi, key=f"server_skala_{i}", label_visibility="collapsed")
            with sc3:
                ans_srv = st.text_area(f"Jwb_C_{i}", key=f"server_ans_{i}", height=68, label_visibility="collapsed", placeholder="Jawaban...")
            
            server_answers[f"q_{i+1}"] = {
                "pertanyaan": q,
                "skala_critical": skala,
                "jawaban": ans_srv
            }
            st.divider()

    # --- TOMBOL SUBMIT ---
    submitted = st.form_submit_button("Submit Assessment", type="primary")

    if submitted:
        if not pic_name:
            st.error("Mohon isi nama PIC Unit Kerja.")
        else:
            payload = {
                "assessment_date": str(tgl_assessment),
                "unit_kerja": unit_kerja,
                "pic_name": pic_name,
                "storage_responses": storage_answers,
                "server_responses": server_answers
            }
            
            try:
                data, count = supabase.table("assessment_bmkg").insert(payload).execute()
                st.success(f"Data Assessment untuk unit {unit_kerja} berhasil disimpan!")
                st.balloons()
                
            except Exception as e:
                st.error(f"Terjadi kesalahan saat menyimpan data: {e}")

# --- TAB 3: HASIL & EXPORT (REDESIGNED) ---
with tab3:
    st.subheader("Daftar Hasil Assessment")
    
    # 1. Search Bar
    search_query = st.text_input("🔍 Cari Unit Kerja atau PIC...", placeholder="Ketik nama unit kerja atau PIC...")
    
    # 2. Fetch Data
    # Note: In a real production app with many records, you should paginate or filter on the server side.
    # For now, we fetch all and filter in Python as requested for simplicity and responsiveness.
    try:
        response = supabase.table("assessment_bmkg").select("*").order("created_at", desc=True).execute()
        all_data = response.data
        
        if all_data:
            # Filter Data
            filtered_data = []
            if search_query:
                query = search_query.lower()
                filtered_data = [
                    item for item in all_data 
                    if query in item['unit_kerja'].lower() or query in item['pic_name'].lower()
                ]
            else:
                filtered_data = all_data
            
            st.write(f"Menampilkan {len(filtered_data)} data assessment.")
            st.markdown("---")
            
            # 3. Display List
            if not filtered_data:
                st.info("Tidak ada data yang cocok dengan pencarian.")
            
            for idx, item in enumerate(filtered_data):
                # Container for each row
                with st.container():
                    c1, c2 = st.columns([4, 1])
                    
                    with c1:
                        st.markdown(f"### {idx + 1}. {item['unit_kerja']}")
                        st.caption(f"PIC: {item['pic_name']} | Tanggal: {item['assessment_date']} | Created: {item['created_at'][:10]}")
                    
                    with c2:
                        # Prepare Data for Excel on the fly (or pre-calculate if needed, but here we do it per item)
                        # Parsing JSON
                        try:
                            storage_res = item['storage_responses']
                            if isinstance(storage_res, str): storage_res = json.loads(storage_res)
                            
                            server_res = item['server_responses']
                            if isinstance(server_res, str): server_res = json.loads(server_res)
                            
                            # Transform to List
                            s_data = []
                            for k, v in storage_res.items():
                                no = k.split('_')[1] if '_' in k else k
                                s_data.append({"No": no, "Pertanyaan": v['pertanyaan'], "Keterangan": v['keterangan'], "Jawaban": v['jawaban']})
                            
                            srv_data = []
                            for k, v in server_res.items():
                                no = k.split('_')[1] if '_' in k else k
                                srv_data.append({"No": no, "Pertanyaan": v['pertanyaan'], "Skala Critical": v['skala_critical'], "Jawaban": v['jawaban']})
                            
                            # Info dict
                            info_dict = {
                                'unit_kerja': item['unit_kerja'],
                                'pic_name': item['pic_name'],
                                'date': item['assessment_date']
                            }
                            
                            # Generate Excel
                            excel_bytes = create_styled_excel(s_data, srv_data, info_dict)
                            
                            st.download_button(
                                label="📥 Download XLSX",
                                data=excel_bytes,
                                file_name=f"Assessment_{item['unit_kerja']}_{item['assessment_date']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"btn_dl_{item['id']}" # Unique key required
                            )
                        except Exception as e:
                            st.error(f"Error generating file: {e}")
                    
                    st.divider()
        else:
            st.info("Belum ada data assessment yang tersimpan di database.")
            
    except Exception as e:
        st.error(f"Gagal mengambil data dari database: {e}")